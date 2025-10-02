import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, Reference, BarChart

DB_PATH = "app.db"

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ---------------- Clients ----------------
def refresh_clients(listbox, search_var=None):
    listbox.delete(0, tk.END)
    conn = get_conn()
    rows = conn.execute("SELECT nom FROM clients ORDER BY nom").fetchall()
    conn.close()
    filtre = (search_var.get().lower() if search_var else "").strip()
    for r in rows:
        if not filtre or filtre in r["nom"].lower():
            listbox.insert(tk.END, r["nom"])

def add_client(listbox, search_var):
    name = simpledialog.askstring("Ajouter client", "Nom du client :")
    if not name:
        return
    conn = get_conn()
    try:
        conn.execute("INSERT INTO clients (nom) VALUES (?)", (name.strip(),))
        conn.commit()
    except sqlite3.IntegrityError:
        messagebox.showerror("Erreur", "Ce client existe déjà.")
    conn.close()
    refresh_clients(listbox, search_var)

def delete_client(listbox, search_var):
    selection = listbox.curselection()
    if not selection:
        return
    name = listbox.get(selection[0])
    if messagebox.askyesno("Supprimer", f"Supprimer le client {name} ?"):
        conn = get_conn()
        conn.execute("DELETE FROM clients WHERE nom=?", (name,))
        conn.commit()
        conn.close()
        refresh_clients(listbox, search_var)

def make_clients_tab(parent):
    frame = ttk.Frame(parent)
    search_var = tk.StringVar()
    search_entry = ttk.Entry(frame, textvariable=search_var)
    search_entry.pack(fill="x", padx=5, pady=5)
    listbox = tk.Listbox(frame, font=("Segoe UI", 11))
    listbox.pack(fill="both", expand=True, padx=5, pady=5)
    btn_add = tk.Button(frame, text="Ajouter client", command=lambda: add_client(listbox, search_var))
    btn_add.pack(pady=5)
    menu = tk.Menu(frame, tearoff=0)
    menu.add_command(label="Supprimer", command=lambda: delete_client(listbox, search_var))
    def show_menu(event):
        if listbox.curselection():
            menu.tk_popup(event.x_root, event.y_root)
    listbox.bind("<Button-3>", show_menu)
    search_var.trace_add("write", lambda *args: refresh_clients(listbox, search_var))
    refresh_clients(listbox, search_var)
    return frame

# ---------------- Employés ----------------
def refresh_employes(listbox, search_var=None):
    listbox.delete(0, tk.END)
    conn = get_conn()
    rows = conn.execute("SELECT nom FROM employes ORDER BY nom").fetchall()
    conn.close()
    filtre = (search_var.get().lower() if search_var else "").strip()
    for r in rows:
        if not filtre or filtre in r["nom"].lower():
            listbox.insert(tk.END, r["nom"])

def add_employe(listbox, search_var):
    name = simpledialog.askstring("Ajouter employé", "Nom de l'employé :")
    if not name:
        return
    conn = get_conn()
    try:
        conn.execute("INSERT INTO employes (nom) VALUES (?)", (name.strip(),))
        conn.commit()
    except sqlite3.IntegrityError:
        messagebox.showerror("Erreur", "Cet employé existe déjà.")
    conn.close()
    refresh_employes(listbox, search_var)

def delete_employe(listbox, search_var):
    selection = listbox.curselection()
    if not selection:
        return
    name = listbox.get(selection[0])
    if messagebox.askyesno("Supprimer", f"Supprimer l'employé {name} ?"):
        conn = get_conn()
        conn.execute("DELETE FROM employes WHERE nom=?", (name,))
        conn.commit()
        conn.close()
        refresh_employes(listbox, search_var)

def make_employes_tab(parent):
    frame = ttk.Frame(parent)
    search_var = tk.StringVar()
    search_entry = ttk.Entry(frame, textvariable=search_var)
    search_entry.pack(fill="x", padx=5, pady=5)
    listbox = tk.Listbox(frame, font=("Segoe UI", 11))
    listbox.pack(fill="both", expand=True, padx=5, pady=5)
    btn_add = tk.Button(frame, text="Ajouter employé", command=lambda: add_employe(listbox, search_var))
    btn_add.pack(pady=5)
    menu = tk.Menu(frame, tearoff=0)
    menu.add_command(label="Supprimer", command=lambda: delete_employe(listbox, search_var))
    def show_menu(event):
        if listbox.curselection():
            menu.tk_popup(event.x_root, event.y_root)
    listbox.bind("<Button-3>", show_menu)
    search_var.trace_add("write", lambda *args: refresh_employes(listbox, search_var))
    refresh_employes(listbox, search_var)
    return frame

# ---------------- Remplissages (tableau croisé + export) ----------------
def refresh_remplissages(tree, search_var=None):
    for row in tree.get_children():
        tree.delete(row)
    conn = get_conn()
    pdfs = [row["pdf_modele"] for row in conn.execute("SELECT DISTINCT pdf_modele FROM remplissages ORDER BY pdf_modele").fetchall()]
    clients = [row["nom"] for row in conn.execute("SELECT nom FROM clients ORDER BY nom").fetchall()]
    data = {}
    rows = conn.execute("SELECT client_nom, pdf_modele, statut, employe_nom, date_remplissage FROM remplissages").fetchall()
    for r in rows:
        data[(r["client_nom"], r["pdf_modele"])] = r["statut"]
    conn.close()
    filtre = (search_var.get().lower() if search_var else "").strip()
    for client in clients:
        values = [client]
        for pdf in pdfs:
            values.append(data.get((client, pdf), ""))
        row_text = " ".join(values).lower()
        if not filtre or filtre in row_text:
            tree.insert("", tk.END, values=values)
    return pdfs, clients, data

def edit_cell(tree, pdfs, search_var):
    item = tree.selection()
    if not item: return
    iid = item[0]
    values = tree.item(iid, "values")
    client_nom = values[0]
    col = tree.identify_column(tree.winfo_pointerx() - tree.winfo_rootx())
    col_index = int(col.replace("#", "")) - 1
    if col_index == 0: return
    pdf_modele = pdfs[col_index - 1]
    new_statut = simpledialog.askstring("Modifier statut", f"Statut pour {client_nom} - {pdf_modele} (i/s/v) :",
                                        initialvalue=values[col_index])
    if not new_statut: return
    if new_statut not in ("i", "s", "v"):
        messagebox.showerror("Erreur", "Statut invalide (i/s/v seulement)")
        return
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM remplissages WHERE client_nom=? AND pdf_modele=?", (client_nom, pdf_modele))
    row = cur.fetchone()
    if row:
        conn.execute("UPDATE remplissages SET statut=? WHERE client_nom=? AND pdf_modele=?",
                     (new_statut, client_nom, pdf_modele))
    else:
        conn.execute("INSERT INTO remplissages (client_nom, pdf_modele, statut) VALUES (?,?,?)",
                     (client_nom, pdf_modele, new_statut))
    conn.commit()
    conn.close()
    refresh_remplissages(tree, search_var)

def export_remplissages_to_excel():
    conn = get_conn()
    pdfs = [row["pdf_modele"] for row in conn.execute("SELECT DISTINCT pdf_modele FROM remplissages ORDER BY pdf_modele").fetchall()]
    clients = [row["nom"] for row in conn.execute("SELECT nom FROM clients ORDER BY nom").fetchall()]
    rows = conn.execute("SELECT client_nom, pdf_modele, statut, employe_nom, date_remplissage FROM remplissages").fetchall()
    conn.close()

    # Construire tableau croisé
    data = {(r["client_nom"], r["pdf_modele"]): r["statut"] for r in rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Remplissages"

    headers = ["Client"] + pdfs
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for client in clients:
        row = [client]
        for pdf in pdfs:
            row.append(data.get((client, pdf), ""))
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Feuille Stats
    ws_stats = wb.create_sheet("Stats")

    # 1️⃣ Répartition statuts
    ws_stats.append(["Statut", "Nombre"])
    stats = {"i": 0, "s": 0, "v": 0}
    for r in rows:
        if r["statut"] in stats:
            stats[r["statut"]] += 1
    for k, v in stats.items():
        ws_stats.append([k, v])
    chart1 = PieChart()
    chart1.title = "Répartition des statuts"
    labels = Reference(ws_stats, min_col=1, min_row=2, max_row=4)
    data_ref = Reference(ws_stats, min_col=2, min_row=1, max_row=4)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(labels)
    ws_stats.add_chart(chart1, "D5")

    # 2️⃣ Répartition par employé
    ws_stats.append([])
    ws_stats.append(["Employé", "Nombre"])
    employe_counts = {}
    for r in rows:
        emp = r["employe_nom"]
        if emp:
            employe_counts[emp] = employe_counts.get(emp, 0) + 1
    for emp, count in employe_counts.items():
        ws_stats.append([emp, count])
    if employe_counts:
        start_row = 7
        end_row = start_row + len(employe_counts)
        chart2 = PieChart()
        chart2.title = "Répartition par employé"
        labels2 = Reference(ws_stats, min_col=1, min_row=start_row+1, max_row=end_row)
        data_ref2 = Reference(ws_stats, min_col=2, min_row=start_row, max_row=end_row)
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(labels2)
        ws_stats.add_chart(chart2, "D20")

    # 3️⃣ Répartition mensuelle par employé
    ws_stats.append([])
    ws_stats.append(["Employé/Mois"])
    # Construire dict {employe: {mois: count}}
    monthly = {}
    for r in rows:
        if not r["employe_nom"] or not r["date_remplissage"]:
            continue
        mois = r["date_remplissage"][:7]  # YYYY-MM
        emp = r["employe_nom"]
        monthly.setdefault(emp, {})
        monthly[emp][mois] = monthly[emp].get(mois, 0) + 1
    mois_list = sorted({m for d in monthly.values() for m in d.keys()})
    ws_stats.cell(row=ws_stats.max_row, column=2, value=" | ".join(mois_list))
    # Écrire tableau
    ws_stats.append(["Employé"] + mois_list)
    for emp, d in monthly.items():
        row = [emp]
        for mois in mois_list:
            row.append(d.get(mois, 0))
        ws_stats.append(row)

    if monthly:
        start_row = ws_stats.max_row - len(monthly)
        end_row = ws_stats.max_row
        start_col = 2
        end_col = 1 + len(mois_list)
        chart3 = BarChart()
        chart3.title = "Remplissages mensuels par employé"
        data_ref3 = Reference(ws_stats, min_col=start_col, min_row=start_row, max_row=end_row)
        cats = Reference(ws_stats, min_col=1, min_row=start_row+1, max_row=end_row)
        chart3.add_data(data_ref3, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.y_axis.title = "Nombre de PDF"
        chart3.x_axis.title = "Employés"
        ws_stats.add_chart(chart3, "H5")

    filename = "remplissages_croise.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export réussi", f"Tableau croisé + graphiques exportés : {filename}")

def make_remplissages_tab(parent):
    frame = ttk.Frame(parent)
    top_frame = ttk.Frame(frame)
    top_frame.pack(fill="x", padx=5, pady=5)
    search_var = tk.StringVar()
    search_entry = ttk.Entry(top_frame, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
    btn_export = tk.Button(top_frame, text="Exporter en Excel", command=export_remplissages_to_excel)
    btn_export.pack(side="right")
    conn = get_conn()
    pdfs = [row["pdf_modele"] for row in conn.execute("SELECT DISTINCT pdf_modele FROM remplissages ORDER BY pdf_modele").fetchall()]
    conn.close()
    cols = ["Client"] + pdfs
    tree = ttk.Treeview(frame, columns=cols, show="headings")
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=120, anchor="center")
    tree.pack(fill="both", expand=True, padx=5, pady=5)
    menu = tk.Menu(frame, tearoff=0)
    menu.add_command(label="Modifier statut", command=lambda: edit_cell(tree, pdfs, search_var))
    def show_context_menu(event):
        iid = tree.identify_row(event.y)
        if iid:
            tree.selection_set(iid)
            menu.tk_popup(event.x_root, event.y_root)
    tree.bind("<Button-3>", show_context_menu)
    search_var.trace_add("write", lambda *args: refresh_remplissages(tree, search_var))
    refresh_remplissages(tree, search_var)
    return frame

# ---------------- Main ----------------
def main():
    root = tk.Tk()
    root.title("Application PC")
    root.geometry("1300x700")
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)
    tab_clients = make_clients_tab(notebook)
    notebook.add(tab_clients, text="Clients")
    tab_employes = make_employes_tab(notebook)
    notebook.add(tab_employes, text="Employés")
    tab_remplissages = make_remplissages_tab(notebook)
    notebook.add(tab_remplissages, text="Remplissages")
    root.mainloop()

if __name__ == "__main__":
    main()
